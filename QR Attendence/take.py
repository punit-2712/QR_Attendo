import cv2
import pyzbar.pyzbar as pyzbar
from pandas import ExcelFile, read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tkinter import messagebox
cap = cv2.VideoCapture(0)
def myFunc():
    import pymysql
    conn = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
    mycursor = conn.cursor()
    mycursor.execute("select Gr_no from Student_details")
    myresults= mycursor.fetchall()
    counter = 0
    flag=1
    grno = str(grn)
    gr_No=""
    Qr_gr=""
    for Gr_no in myresults:
        Gr_no = str(Gr_no)
        for i in Gr_no:
            if(i not in ['(',',',')']):                                 #And in database GR NUMBER is the form of (917,)
                gr_No+=i
        for j in grno:
            if (j not in ['b', "'", "'"]):                              #In Qr code, data is in the form of (b'917')

                Qr_gr+=j

        if str(gr_No) == str(Qr_gr):                                    #Comparing Qr-code and database data
            flag=0
            #print("done")
            def make():
                subject="Sheet1"
                GR = [Qr_gr]                                          # Storing all the Gr numbers in the list
                xls = ExcelFile('attendance1.xlsx')
                df = read_excel(xls, "Sheet1")                        #Storing whole excel file in variable df
                today = dt.datetime.now()                             # Current date and time
                today = today.strftime('%d/%m/%y')                    # Current time in string form
                student_status = ['AB' for i in range(len(df))]
                df[today] = student_status                            # Marking AB initially to all student
                excelbook = load_workbook('attendance1.xlsx')         # Will loading in the same file and sheet
                sheet = excelbook.get_sheet_by_name("Sheet1")
                max_col_no = sheet.max_column
                max_col_letter = get_column_letter(sheet.max_column+ 1)
                row_1 = []
                for i in range(max_col_no):
                    row_1.append(sheet.cell(row=1, column=i + 1).value)
                # mark attendance
                print (df['Gr_no'])
                present_status = df[today].values.tolist()
                present_status.insert(0, today)
                                                                    # writing to excel file
                for i in range(len(df) + 1):
                    c = sheet[max_col_letter + str(i + 1)]
                    c.value = present_status[i]
                excelbook.save('attendance1.xlsx')
            make()
            messagebox.showinfo('Done',
                                'Today''s attendance sheet has been successfully created')
            break
        else:
            counter += 1
        gr_No=""
        Qr_gr=""
    if flag==1:
        print(type(Gr_no),"Gr_no")
        print(type(grno),"grn")

while True:
    _, frame = cap.read()
    decodedObjects = pyzbar.decode(frame)
    for obj in decodedObjects:
        print("Data", obj.data)
        grn = obj.data
        break
    cv2.imshow("Frame", frame)
    key = cv2.waitKey(1)
    if key == 27:
        myFunc()
        break










