from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import qrcode
import cv2
import pyzbar.pyzbar as pyzbar
from pandas import ExcelFile, read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pymysql
Qr_gr = ""
def make():

    xls = ExcelFile('attendance1.xlsx')
    df = read_excel(xls, "Sheet1")                                              #Storing whole excel file in variable df
    today = dt.datetime.now()                                                   #Current date and time
    today = today.strftime('%d/%m/%y')                                          #Current time in string form
    student_status = ['AB' for i in range(len(df))]
    df[today] = student_status                                                  #Marking AB initially to all student
    excelbook = load_workbook('attendance1.xlsx')                               #Will loading in the same file and sheet
    sheet = excelbook.get_sheet_by_name("Sheet1")
    max_col_no = sheet.max_column
    max_col_letter = get_column_letter(sheet.max_column + 1)
    row_1 = []
    for i in range(max_col_no):
        row_1.append(sheet.cell(row=1, column=i + 1).value)
                                                                                #mark attendance
    present_status = df[today].values.tolist()
    present_status.insert(0, today)
                                                                                #writing to excel file
    for i in range(len(df) + 1):
        c = sheet[max_col_letter + str(i + 1)]
        c.value = present_status[i]
    excelbook.save('attendance1.xlsx')
    makebtn.config(state=DISABLED)

()

def new():
    gro=Entry3.get()
    name=Entry4.get()                                                                    #Getting entries data from GUI
    std=Entry5.get()
    roll=Entry6.get()
    mydb = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
    mycursor = mydb.cursor()                                                             #Connecting the database
    sql = "INSERT INTO Student_details (Gr_no, name, std, Roll_no) VALUES (%s, %s,%s, %s )"
    val = (gro, name, std, roll)
    mycursor.execute(sql, val)
    QR_code = Entry3.get()  # getting data from GUI
    filename = QR_code + '.jpg'
    img = qrcode.make(QR_code)  # .make makes Qr code
    img.save(filename)  # saving the Qr code
    root.photo = PhotoImage(file=filename)                                          # getting Qrcode Image
    QR.config(image=root.photo, width=300, height=300)
    messagebox.showinfo('Saved', 'QR code saved as " ' + QR_code + ' " successfully!\n\tin current location \n\t And all detials are added to database')
    mydb.commit()
    Entry3.delete(0, "end")
    Entry4.delete(0, "end")                                                              #Reseting the entry after clicking button
    Entry5.delete(0, "end")
    Entry6.delete(0, "end")
()

cap = cv2.VideoCapture(0)

def check():
    def myFunc():
        conn = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
        mycursor = conn.cursor()
        mycursor.execute("select Gr_no from Student_details")
        myresults = mycursor.fetchall()
        counter = 0
        counter2 = 1
        grno = str(grn)                                                     #grn is the data from QR code
        gr_No = ""
        Qr_gr = ""
        for Gr_no in myresults:
            Gr_no = str(Gr_no)
            for i in Gr_no:
                if (i not in ['(', ',', ')']):                              #And in database GR NUMBER is the form of (917,)
                    gr_No += i

            for j in grno:
                if (j not in ['b', "'", "'"]):                              #In Qr code, data was in the form of (b'917')
                    Qr_gr += j


            if str(gr_No) == str(Qr_gr):                                    #Comparing Qr-code and database data
                counter2 = 0


                def save_attendance():

                    xls = ExcelFile('attendance1.xlsx')
                    df = read_excel(xls, "Sheet1")                            #Storing whole excel file in variable df
                    today = dt.datetime.now()                                 # Current date and time
                    today = today.strftime('%d/%m/%y')                        # Current time in string form
                    excelbook = load_workbook('attendance1.xlsx')             # Will loading in the same file and sheet
                    sheet = excelbook.get_sheet_by_name("Sheet1")             # It will return a worksheet of this name
                    max_col_no = sheet.max_column                             # This will create max col size to look pretty.xD
                    max_col_letter = get_column_letter(sheet.max_column)      # this will give letter of the cell address
                    row_1 = []                                                # this is the list of all the titles of the of the sheet
                    for i in range(max_col_no):
                        row_1.append(sheet.cell(row=1, column=i + 1).value)   #appending values of row 1 in row1 list

                                                                              # mark attendance
                    for i in range(0, len(df['Gr_no'])):                      # length is 5 because there are 5 five records only in the worksheet.xD
                        if str(df['Gr_no'][i]) == str(Qr_gr):
                            df[today][i] = "P"                               #Doing Present to that GR NUMBER

                    present_status = df[today].values.tolist()               #converts all the values of the today to list

                    present_status.insert(0, today)                          # in 0th location of today

                                                                              # writing to excel file
                    for i in range(len(df) + 1):
                        c = sheet[max_col_letter + str(i + 1)]               #This is the adresss of the today's attedance
                        c.value = present_status[i]

                    excelbook.save('attendance1.xlsx')
                    messagebox.showinfo('Done ',
                                        'Your today''s attendance has been updated.')

                save_attendance()
                break
            else:
                counter += 1
            gr_No = ""
            Qr_gr = ""
        if counter2 == 1:
            messagebox.showinfo('Error ',
                                'Please Scan registered Gr Number.')

    while True:
        _, frame = cap.read()                                                 #This function will read the frame
        decodedObjects = pyzbar.decode(frame)                                 #This fuction will decode the objects from frame
        for obj in decodedObjects:                                            #To obtain the data from OR code
            print("Data", obj.data)
            grn = obj.data

            break
        cv2.imshow("Frame", frame)
        key = cv2.waitKey(1)                                                  #will display the window infinitely until any keypress
        if key == 27:                                                         #27 is Esc key
            myFunc()
            break
    ()
()


#GUI CODING
root=Tk()
root.title("Attendence System")
root.resizable(width=FALSE, height=FALSE)
root.geometry('1300x1200')
tab_parent= ttk.Notebook(root)
tab1 = Frame(tab_parent)
tab2 = Frame(tab_parent)
tab3 = Frame(tab_parent)
tab_parent.add(tab1, text='Home')
tab_parent.add(tab2, text='Gr code generation')
tab_parent.add(tab3, text='New Record')
tab_parent.pack(expand=1, fill='both')
title = Label(tab1,text = "AIRPORT SCHOOL \n AHMEDABAD", font=('Old English Text MT',30), fg='Brown')
title.place(x=150, y=10)
system = Label(tab1,text = "QR Code Attendance system", font=('Arial',11, 'bold', 'underline'), fg='black')
system.place(x=570, y=440)
punch = Button(tab1, text="Punch Your ID-Card", bg='green', fg='white', width=30, height=8, command = check)
punch.place(x=400, y=170)
makebtn=Button(tab1, text="Create Attedance", bg='green', fg='white', width=30, height=8, command=make)
makebtn.place(x=50, y=170)
QR_code = Label(tab2, text="Gr Number:", font=('Arial', 12,'bold'))
QR_code.place(x=75, y=67)

a = Label(tab3,text = "Gr Number:" , font=('Arial',12, 'bold'), fg='black')
a.place(x=120, y=66)
Entry3 = Entry(tab3, fg='blue', bd=3, width=40)
Entry3.place(x=230, y=70)
b = Label(tab3,text = "Name:", font=('Arial',12,'bold'), fg='black')
b.place(x=120, y=100)
Entry4 = Entry(tab3, fg='blue', bd=3, width=40)
Entry4.place(x=230, y=100)
c = Label(tab3,text = "Std:" , font=('Arial',12,'bold'), fg='black')
c.place(x=120, y=130)
Entry5 = Entry(tab3, fg='blue', bd=3, width=40)
Entry5.place(x=230, y=130)
d = Label(tab3,text = "Roll Number:", font=('Arial',12,'bold'), fg='black')
d.place(x=120, y=160)
Entry6 = Entry(tab3, fg='blue', bd=3, width=40)
Entry6.place(x=230, y=160)
QR = Label(tab3, image='')
QR.place(x=100, y=300)
btn = Button(tab3,text="Submit",bg='green', fg='white', command=new)
btn.place(x=230, y=200)
root.mainloop()



