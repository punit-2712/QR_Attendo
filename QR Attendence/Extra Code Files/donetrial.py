from PIL import Image, ImageDraw, ImageFont
from PIL import Image
image = Image.new('RGB', (1000,900), (255, 255, 255))
draw = ImageDraw.Draw(image)
font = ImageFont.truetype('arial.ttf', size=18)
font1 = ImageFont.truetype('arial.ttf', size=15)

from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import qrcode
import cv2
import pyzbar.pyzbar as pyzbar
import numpy as np
from pandas import ExcelFile, read_excel
import datetime as dt

import os
import cv2
import pywhatkit
from tkinter import *
from tkinter import ttk
import time
import pyzbar.pyzbar as pyzbar
from pandas import ExcelFile, read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pymysql
from pynput.keyboard import Key, Controller
keyboard = Controller()

from pyzbar.pyzbar import decode
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pymysql
import os
from pynput.keyboard import Key, Controller
keyboard = Controller()



def modify():
    start = Entryroll.get()
    mydb = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
    mycursor = mydb.cursor()  # Connecting the database
    '''sql = "SELECT Gr_no, name, std, Roll_no, DOB, Address, Bloodgr, MobileNo from Student_details where Gr_no='914'"
    #mycursor.execute(sql)
    var =mycursor.execute(sql)
    var.fetchall()
    print(var.fetchall()'''
    sql1="SELECT * FROM Student_details WHERE Gr_no=%s"
    val2=(start)
    mycursor.execute(sql1,val2)

    myresult = mycursor.fetchall()
    Label1 = Label(tab2, text="Gr Number", width=10)
    Label1.grid(row=0, column=0)
    Label2 = Label(tab2, text="FULL NAME", width=10)
    Label2.grid(row=0, column=1)
    Label3 = Label(tab2, text="STD", width=10)
    Label3.grid(row=0, column=2)
    Label4 = Label(tab2, text="ROLL NO.", width=10)
    Label4.grid(row=0, column=3)
    Label5 = Label(tab2, text="DOB", width=10)
    Label5.grid(row=0, column=4)
    Label6 = Label(tab2, text="Address", width=30)
    Label6.grid(row=0, column=5)
    Label7 = Label(tab2, text="Blood gr.", width=10)
    Label7.grid(row=0, column=6)
    Label8 = Label(tab2, text="Mobile No.", width=10)
    Label8.grid(row=0, column=7)

    for index, dat in enumerate(myresult):
        Label(tab2, text=dat[0]).grid(row=index + 1, column=0)
        Label(tab2, text=dat[1]).grid(row=index + 1, column=1)
        Label(tab2, text=dat[2]).grid(row=index + 1, column=2)
        Label(tab2, text=dat[3]).grid(row=index + 1, column=3)
        Label(tab2, text=dat[4]).grid(row=index + 1, column=4)
        Label(tab2, text=dat[5]).grid(row=index + 1, column=5)
        Label(tab2, text=dat[6]).grid(row=index + 1, column=6)
        Label(tab2, text=dat[7]).grid(row=index + 1, column=7)
        Edit = Button(tab2, text="EDIT", bg='green', fg='white', command=ED)
        Edit.place(x=20, y=60)

def Upt():
    gro = Entry9.get()
    name1 = Entry4.get()  # Getting entries data from GUI
    std1 = Entry5.get()
    roll = Entry6.get()
    dob = Entry7.get()
    bloodgr = Entry8.get()
    add = Entry9.get()
    mobile = Entry10.get()
    start1 = Entryroll.get()

    mydb = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
    mycursor = mydb.cursor()  # Connecting the database
    sql2="UPDATE Student_details SET Gr_no=%s , name=%s , std=%s , Roll_no=%s , DOB=%s , Address=%s , Bloodgr=%s, MobileNo=%s WHERE Gr_no=%s"
    val3=(gro, name1, std1, roll, dob, bloodgr, add, mobile, start1)
    mycursor.execute(sql2, val3)
    mydb.commit()
()
def new():
    gro=Entry3.get()
    name=Entry4.get()                                                                    #Getting entries data from GUI
    std=Entry5.get()
    roll=Entry6.get()
    dob=Entry7.get()
    bloodgr=Entry8.get()
    add=Entry9.get()
    mobile=Entry10.get()
    mydb = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")
    mycursor = mydb.cursor()                                                             #Connecting the database
    sql = "INSERT INTO Student_details (Gr_no, name, std, Roll_no, DOB, Address, Bloodgr, MobileNo) VALUES (%s, %s,%s, %s, %s, %s,%s, %s )"
    val = (gro, name, std, roll, dob, add, bloodgr, mobile)
    mycursor.execute(sql, val)
    QR_code = Entry3.get()  # getting data from GUI
    filename = QR_code + '.jpg'
    img = qrcode.make(QR_code)  # .make makes Qr code
    img.save(filename)  # saving the Qr code
    root.photo = PhotoImage(file=filename)                                          # getting Qrcode Image
    QR.config(image=root.photo, width=300, height=300)
    #messagebox.showinfo('Saved', 'QR code saved as " ' + QR_code + ' " successfully!\n\tin current location \n\t And all detials are added to database')
    mydb.commit()
    from PIL import Image , ImageDraw, ImageFont
    im = Image.open("ID.JPG")

    d = ImageDraw.Draw(im)
    location = (320, 125)
    text_color = (255, 255, 255)
    d.text(location, name , font=font,  fill=text_color)

    location = (405, 173)
    text_color = (0, 0, 0)
    d.text(location, QR_code , font=font, fill=text_color)

    location = (270, 220)
    text_color = (0, 0, 0)
    d.text(location, "STD: " + std, font=font1, fill=text_color)

    location = (270, 250)
    text_color = (0, 0, 0)
    d.text(location, "DOB: " + dob, font=font1, fill=text_color)


    location = (270, 280)
    text_color = (0, 0, 0)
    d.text(location, "MOBILE: " + mobile, font=font1, fill=text_color)

    location = (270, 310)
    text_color = (0, 0, 0)
    d.text(location, "ADDRESS: " + add, font=font1, fill=text_color)
    path='C:/Users/VK/PycharmProjects/QR Attendence/ID CARDS/'
    im.save('hello.png')
    til = Image.open(r"hello.png")
    im1 = Image.open(filename)  # 25x25
    images=im1.resize((160,160))
    til.paste(images, (60, 180))
    til.save(name + ".png")



    Entry3.delete(0, "end")
    Entry4.delete(0, "end")                                                              #Reseting the entry after clicking button
    Entry5.delete(0, "end")
    Entry6.delete(0, "end")
    Entry7.delete(0, "end")
    Entry8.delete(0, "end")
    Entry9.delete(0, "end")
    Entry10.delete(0, "end")
()



def ED():
    gr = Label(tab2, text="Gr Number:", font=('Arial', 12, 'bold'), fg='black')
    gr.place(x=150, y=180)
    Entry9 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry9.place(x=250, y=180)

    b = Label(tab2, text="Name:", font=('Arial', 12, 'bold'), fg='black')
    b.place(x=150, y=215)
    Entry4 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry4.place(x=250, y=215)

    c = Label(tab2, text="Std:", font=('Arial', 12, 'bold'), fg='black')
    c.place(x=150, y=250)
    Entry5 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry5.place(x=250, y=250)

    d = Label(tab2, text="Roll Number:", font=('Arial', 12, 'bold'), fg='black')
    d.place(x=150, y=280)
    Entry6 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry6.place(x=250, y=280)

    e = Label(tab2, text="DOB", font=('Arial', 12, 'bold'), fg='black')
    e.place(x=150, y=320)
    Entry7 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry7.place(x=250, y=320)

    f = Label(tab2, text="Blood Group:", font=('Arial', 12, 'bold'), fg='black')
    f.place(x=150, y=350)
    Entry8 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry8.place(x=250, y=350)

    h = Label(tab2, text="Mobile No:", font=('Arial', 12, 'bold'), fg='black')
    h.place(x=150, y=385)
    Entry10 = Entry(tab2, fg='blue', bd=3, width=40)
    Entry10.place(x=250, y=385)

    donegr = Label(tab2, text="Gr Number:", font=('Arial', 12, 'bold'), fg='black')
    a.place(x=120, y=66)
    Entryroll = Entry(tab2, fg='blue', bd=3, width=40)
    Entryroll.place(x=250, y=70)
    upd = Button(tab2, text="Update", bg='green', fg='white', command=Upt)
    upd.place(x=350, y=420)


()

cap=cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)
def check():
    import trial

    def myfunc():

        

        '''file1 = open("Myfile.txt", "r")
        line = file1.readline()
        xls = ExcelFile('attendance2.xlsx')
        df = read_excel(xls, "Sheet1")  # Storing whole excel file in variable df
        today = dt.datetime.now()  # Current date and time
        today = today.strftime('%d/%m/%y')  # Current time in string form
        present_status = df[today].values.tolist()
        print(present_status)
        excelbook = load_workbook('attendance2.xlsx')  # Will loading in the same file and sheet
        sheet = excelbook.get_sheet_by_name("Sheet1")
        present_status.insert(0, today)
        max_col_no = sheet.max_column
        max_col_letter = get_column_letter(sheet.max_column + 1)

        lister = []

        while line:

            line = line.rstrip("\n")
            line = int(line)

            if line not in lister:
                lister.append(line)
                print('here starts')

                line = int(line)

            for i in range(0, len(df['Gr_no'])):

                if df['Gr_no'][i] == line:
                    present_status[i] = 'P'
                    print(present_status[i])
                    for i in range(len(df) + 1):
                        c = sheet[max_col_letter + str(i + 1)]  # This is the adresss of the today's attedance
                        c.value = present_status[i]

                    excelbook.save('attendance2.xlsx')
                    #pywhatkit.sendwhatmsg("+917567091505", "Airport School Alert: Dear Parents, Your ward was abseent on"+today+"lectures.", 12, 52)

                    print("DONE")
                    print(lister)
            line = file1.readline()
    while True:
        sucess, img = cap.read()
        for barcode in decode(img):
            myData = barcode.data.decode('utf-8')
            pts = np.array([barcode.polygon], np.int32)
            pts = pts.reshape((-1, 1, 2))
            cv2.polylines(img, [pts], True, (255, 0, 255), 5)
            pts2 = barcode.rect
            cv2.putText(img, myData, (pts2[0], pts2[1]), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 255))

            file1 = open("Myfile.txt", "a")
            file1.truncate()
            file1.write(myData + '\n')

            file1.close()
        cv2.imshow('result', img)

        key = cv2.waitKey(1)
        if key == 27:
            myfunc()
            break'''


()
#def make():

#()

# GUI CODING
root = Tk()
root.title("Attendence System")
root.resizable(width=FALSE, height=FALSE)
root.geometry('1300x1200')
tab_parent = ttk.Notebook(root)
tab1 = Frame(tab_parent)
tab2 = Frame(tab_parent)
tab3 = Frame(tab_parent)
tab_parent.add(tab1, text='Home')
tab_parent.add(tab2, text='Modification')
tab_parent.add(tab3, text='New Entry')
tab_parent.pack(expand=1, fill='both')
title = Label(tab1, text="AIRPORT SCHOOL \n AHMEDABAD", font=('Old English Text MT', 30), fg='Brown')
title.place(x=150, y=10)
system = Label(tab1, text="QR Code Attendance system", font=('Arial', 11, 'bold', 'underline'), fg='black')
system.place(x=570, y=440)
punch = Button(tab1, text="Punch Your ID-Card", bg='green', fg='white', width=30, height=8, command=check)
punch.place(x=400, y=170)
makebtn = Button(tab1, text="Create Attedance", bg='green', fg='white', width=30, height=8)
makebtn.place(x=50, y=170)
QR_code = Label(tab2, text="Gr Number:", font=('Arial', 12, 'bold'))
QR_code.place(x=75, y=67)

a = Label(tab3, text="Gr Number:", font=('Arial', 12, 'bold'), fg='black')
a.place(x=120, y=66)
Entry3 = Entry(tab3, fg='blue', bd=3, width=40)
Entry3.place(x=250, y=70)
b = Label(tab3, text="Name:", font=('Arial', 12, 'bold'), fg='black')
b.place(x=120, y=100)
Entry4 = Entry(tab3, fg='blue', bd=3, width=40)
Entry4.place(x=250, y=100)
c = Label(tab3, text="Std:", font=('Arial', 12, 'bold'), fg='black')
c.place(x=120, y=130)
Entry5 = Entry(tab3, fg='blue', bd=3, width=40)
Entry5.place(x=250, y=130)
d = Label(tab3, text="Roll Number:", font=('Arial', 12, 'bold'), fg='black')
d.place(x=120, y=160)
Entry6 = Entry(tab3, fg='blue', bd=3, width=40)
Entry6.place(x=250, y=160)

e = Label(tab3, text="DOB", font=('Arial', 12, 'bold'), fg='black')
e.place(x=120, y=190)
Entry7 = Entry(tab3, fg='blue', bd=3, width=40)
Entry7.place(x=250, y=190)

f = Label(tab3, text="Blood Group:", font=('Arial', 12, 'bold'), fg='black')
f.place(x=120, y=220)
Entry8 = Entry(tab3, fg='blue', bd=3, width=40)
Entry8.place(x=250, y=220)

g = Label(tab3, text="Address:", font=('Arial', 12, 'bold'), fg='black')
g.place(x=120, y=250)
Entry9 = Entry(tab3, fg='blue', bd=3, width=40)
Entry9.place(x=250, y=250)

h = Label(tab3, text="Mobile Number:", font=('Arial', 12, 'bold'), fg='black')
h.place(x=120, y=280)
Entry10 = Entry(tab3, fg='blue', bd=3, width=40)
Entry10.place(x=250, y=280)


QR = Label(tab3, image='')
QR.place(x=100, y=350)
btn = Button(tab3, text="Submit", bg='green', fg='white', command=new)
btn.place(x=330, y=310)

donegr = Label(tab2, text="Gr Number:", font=('Arial', 12, 'bold'), fg='black')
a.place(x=120, y=66)
Entryroll = Entry(tab2, fg='blue', bd=3, width=40)
Entryroll.place(x=250, y=70)
btn5 = Button(tab2, text="Submit", bg='green', fg='white', command=modify)
btn5.place(x=330, y=100)
root.mainloop()



