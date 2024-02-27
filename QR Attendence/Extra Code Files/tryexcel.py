import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
from pandas import ExcelFile, read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import datetime
import time


cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_PLAIN

show=0
def myFunc():
    import pymysql

    conn = pymysql.connect(host="localhost", user="root", passwd="punit", db="Project")


    mycursor = conn.cursor()
    mycursor.execute("select Gr_no from Student_details")

    myresults= mycursor.fetchall()

    counter = 0
    flag=1
    grno = str(grn)
    gr_No=""
    Qr_gr=""

    for Gr_no in myresults:
        Gr_no = str(Gr_no)
        for i in Gr_no:
            if(i not in ['(',',',')']):
                gr_No+=i
        for j in grno:
            if (j not in ['b', "'", "'"]):

                Qr_gr+=j

        df = pd.read_csv("Student_details.csv")
        col_names = ['Gr_no', 'name', 'Date', 'Time']
        attendance = pd.DataFrame(columns=col_names)
        if str(gr_No) == str(Qr_gr):
            flag=0
            print("done")
            def save_attendance(name):

                subject="Sheet1"
                names=set(name)
                print(names)
                # df=read_csv('attendance_data/attendance.csv',index_col='Name')
                xls = ExcelFile('attendance1.xlsx')
                df = read_excel(xls, "Sheet1")

                today = dt.datetime.now()
                today = today.strftime('%d/%m/%y')

                '''student_status = ['AB' for i in range(len(df))]
                #print(df["name"])
                df[today] = student_status  # Marking NIL initially to all student'''



                excelbook = load_workbook('attendance1.xlsx')
                sheet = excelbook.get_sheet_by_name("Sheet1")
                max_col_no = sheet.max_column
                max_col_letter = get_column_letter(sheet.max_column)
                row_1 = []

                for i in range(max_col_no):
                    row_1.append(sheet.cell(row=1, column=i + 1).value)

                """if today in row_1:
                    print('Attendance already marked for today !!')
                    return"""
                #excelbook.save('temp.xlsx')
                # mark attendance

                print (df['Gr_no'])
                for i in range(0,len(df['Gr_no'])):
                    if str(df['Gr_no'][i])==str(Qr_gr):
                        print ("OKKK")
                        df[today][i]="P"
                    else:
                        print("NOKKK")

                    #df.loc[i][today] = 'Yes'

                present_status = df[today].values.tolist()
                present_status.insert(0, today)

                # writing to excel file
                for i in range(len(df) + 1):
                    # print(max_col_letter+str(i+1))  #for debugging
                    c = sheet[max_col_letter + str(i + 1)]
                    c.value = present_status[i]
                excelbook.save('attendance1.xlsx')
            save_attendance(name=[Qr_gr])
            break
        else:
            counter += 1
        gr_No=""
        Qr_gr=""
    if flag==1:
        print(type(Gr_no),"Gr_no")
        print(type(grno),"grn")






while True:
    _, frame = cap.read()

    decodedObjects = pyzbar.decode(frame)
    for obj in decodedObjects:
        print("Data", obj.data)
        grn = obj.data

        break
        print("hello")

        #cv2.putText(frame, str(obj.data), (50, 50), font, 2,
                   # (255, 0, 0), 3)

    cv2.imshow("Frame", frame)



    key = cv2.waitKey(1)
    if key == 27:
        myFunc()
        break
def makeA(name):
    Qr_gr=myFunc()
    subject = "Sheet1"
    names = set(name)
    print(names)
    xls = ExcelFile('Student_details.xlsx')
    df = read_excel(xls, "Sheet1")

    today = dt.datetime.now()
    today = today.strftime('%d/%m/%y')

    student_status = ['AB' for i in range(len(df))]
    # print(df["name"])
    df[today] = student_status  # Marking NIL initially to all student

    excelbook = load_workbook('Student_details.xlsx')
    sheet = excelbook.get_sheet_by_name("Sheet1")
    max_col_no = sheet.max_column
    max_col_letter = get_column_letter(sheet.max_column + 1)
    row_1 = []

    for i in range(max_col_no):
        row_1.append(sheet.cell(row=1, column=i + 1).value)


    # mark attendance
    excelbook.save('attendance1.xlsx')
makeA(name=[Qr_gr])
