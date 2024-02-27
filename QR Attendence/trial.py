import os
import cv2
from pandas import ExcelFile, read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tkinter import messagebox
import pymysql
from pynput.keyboard import Key, Controller
keyboard = Controller()
import numpy as np
from pyzbar.pyzbar import decode
cap=cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)

xls = ExcelFile('attendance2.xlsx')
df = read_excel(xls, "Sheet1")  # Storing whole excel file in variable df
today = dt.datetime.now()  # Current date and time
today = today.strftime('%d/%m/%y')  # Current time in string form
student_status = ['AB' for i in range(len(df))]
df[today] = student_status  # Marking AB initially to all student
excelbook = load_workbook('attendance2.xlsx')  # Will loading in the same file and sheet
sheet = excelbook.get_sheet_by_name("Sheet1")
max_col_no = sheet.max_column
max_col_letter = get_column_letter(sheet.max_column + 1)
row_1 = []
for i in range(max_col_no):
    row_1.append(sheet.cell(row=1, column=i + 1).value)
    # mark attendance'''
present_status = df[today].values.tolist()
present_status.insert(0, today)

for i in range(len(df) + 1):
    c = sheet[max_col_letter + str(i + 1)]
    c.value = present_status[i]
excelbook.save('attendance2.xlsx')



os.remove("Myfile.txt")

def myfunc():
    file1 = open("Myfile.txt", "r")
    line = file1.readline()
    xls = ExcelFile('attendance2.xlsx')
    df = read_excel(xls, "Sheet1")
    lister=[]
    while line:
        line=line.rstrip("\n")
        line = int(line)
        if line not in lister:
            lister.append(line)
            line = int(line)
        for i in range(0, len(df['Gr_no'])):
                if df['Gr_no'][i] == line:
                    present_status[i+1]='P'
                    #print(present_status[i])
                    for i in range(len(df) + 1):
                        c = sheet[max_col_letter + str(i + 1)]               #This is the adresss of the today's attedance
                        c.value = present_status[i]
                    excelbook.save('attendance2.xlsx')
        line=file1.readline()

while True:
    sucess, img= cap.read()
    for barcode in decode(img):
        myData=barcode.data.decode('utf-8')
        B=int(myData)
        pts = np.array([barcode.polygon], np.int32)
        pts = pts.reshape((-1, 1, 2))
        cv2.polylines(img, [pts], True, (255, 0, 255), 2)
        pts2 = barcode.rect
        for i in range(0, len(df['Gr_no'])):
                if df['Gr_no'][i] == B:
                    xls = ExcelFile('attendance2.xlsx')
                    df = read_excel(xls, "Sheet1")
                    A=df['name'][i]

        cv2.putText(img, myData +" "+ A +" is present today.", (pts2[0], pts2[1]), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0))
        file1 = open("Myfile.txt", "a")
        file1.write(myData+'\n')
        file1.close()
    cv2.imshow('Scan here',img)
    key = cv2.waitKey(1)
    if key == 27:
        myfunc()
        break
messagebox.showinfo('Done ',"Your class's today's attendance has been updated.")
